import csv
import json
import zipfile
from io import BytesIO, StringIO
from pathlib import Path
from typing import Annotated, Any, Literal
from urllib.parse import quote
from xml.etree import ElementTree

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.auth import require_admin
from app.database import get_db
from app.models import Character, Work
from app.schemas.imports import MetadataImportRow, MetadataImportSummary

router = APIRouter(prefix="/imports", tags=["imports"])

WORK_FIELDS = {
    "name",
    "original_name",
    "aliases",
    "description",
    "tagline",
    "production_year",
    "run_time_minutes",
    "community_rating",
    "content_rating",
    "genres",
    "studios",
    "official_site",
    "status",
    "sort_order",
}
CHARACTER_FIELDS = {"name", "original_name", "aliases", "description"}
WORK_NAME_KEYS = ("work_name", "作品", "作品名", "name")
CHARACTER_NAME_KEYS = ("character_name", "角色", "角色名")
TEMPLATE_FORMATS = {"csv", "json", "xlsx", "xlsm"}
TEMPLATE_HEADERS = [
    "work_name",
    "work_original_name",
    "work_aliases",
    "work_description",
    "work_tagline",
    "work_production_year",
    "work_run_time_minutes",
    "work_community_rating",
    "work_content_rating",
    "work_genres",
    "work_studios",
    "work_official_site",
    "work_status",
    "work_sort_order",
    "character_name",
    "character_original_name",
    "character_aliases",
    "character_description",
]
TEMPLATE_ROWS = [
    {
        "work_name": "示例作品A",
        "work_original_name": "Example Work A",
        "work_aliases": "示例别名A; Example Alias A",
        "work_description": "这里只填写作品资料，不创建角色。",
        "work_tagline": "替换为作品标语",
        "work_production_year": "2024",
        "work_run_time_minutes": "24",
        "work_community_rating": "8.7",
        "work_content_rating": "safe",
        "work_genres": "校园; 推理",
        "work_studios": "示例动画公司",
        "work_official_site": "https://example.com/work-a",
        "work_status": "完结",
        "work_sort_order": "10",
        "character_name": "",
        "character_original_name": "",
        "character_aliases": "",
        "character_description": "",
    },
    {
        "work_name": "示例作品B",
        "work_original_name": "Example Work B",
        "work_aliases": "示例别名B",
        "work_description": "这一行会创建或更新作品，并创建或更新作品下的角色。",
        "work_tagline": "替换为作品标语",
        "work_production_year": "2025",
        "work_run_time_minutes": "25",
        "work_community_rating": "9.1",
        "work_content_rating": "sensitive",
        "work_genres": "奇幻; 冒险",
        "work_studios": "示例制作组",
        "work_official_site": "https://example.com/work-b",
        "work_status": "连载",
        "work_sort_order": "20",
        "character_name": "示例角色B",
        "character_original_name": "Example Character B",
        "character_aliases": "角色别名B",
        "character_description": "替换为角色简介。",
    },
]
TEMPLATE_MEDIA_TYPES = {
    "csv": "text/csv; charset=utf-8",
    "json": "application/json; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "xlsm": "application/vnd.ms-excel.sheet.macroEnabled.12",
}
CONTENT_TYPES_PATH = "[Content_Types].xml"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
XLSM_WORKBOOK_CONTENT_TYPE = "application/vnd.ms-excel.sheet.macroEnabled.main+xml"


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _number(value: Any, kind: type[int] | type[float]) -> int | float | None:
    text = _clean(value)
    if text is None:
        return None
    try:
        return kind(text)
    except ValueError as exc:
        raise ValueError(f"{text} is not a valid {kind.__name__}") from exc


def _pick(row: dict[str, Any], keys: tuple[str, ...]) -> str | None:
    for key in keys:
        if key in row:
            value = _clean(row.get(key))
            if value:
                return value
    return None


def _normalize_row(raw: dict[str, Any], row_number: int) -> dict[str, Any]:
    normalized = {str(key).strip(): value for key, value in raw.items() if key is not None}
    work_name = _pick(normalized, WORK_NAME_KEYS)
    character_name = _pick(normalized, CHARACTER_NAME_KEYS)
    result: dict[str, Any] = {
        "row_number": row_number,
        "work_name": work_name,
        "character_name": character_name,
        "work": {},
        "character": {},
    }
    for field in WORK_FIELDS:
        value = _clean(normalized.get(f"work_{field}", normalized.get(field)))
        if value is not None:
            result["work"][field] = value
    for field in CHARACTER_FIELDS:
        value = _clean(normalized.get(f"character_{field}"))
        if value is not None:
            result["character"][field] = value
    if "name" in result["work"]:
        result["work_name"] = result["work"].pop("name")
    if character_name:
        result["character"]["name"] = character_name
    for int_field in ("production_year", "run_time_minutes", "sort_order"):
        if int_field in result["work"]:
            result["work"][int_field] = _number(result["work"][int_field], int)
    if "community_rating" in result["work"]:
        result["work"]["community_rating"] = _number(result["work"]["community_rating"], float)
    return result


def _parse_csv(data: bytes) -> list[dict[str, Any]]:
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    return [dict(row) for row in reader]


def _parse_json(data: bytes) -> list[dict[str, Any]]:
    payload = json.loads(data.decode("utf-8-sig"))
    if isinstance(payload, dict):
        payload = payload.get("items") or payload.get("rows") or []
    if not isinstance(payload, list):
        raise ValueError("JSON must be an array or an object with items/rows")
    return [dict(item) for item in payload if isinstance(item, dict)]


def _parse_xlsx(data: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Excel import requires openpyxl to be installed") from exc
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    items: list[dict[str, Any]] = []
    for row in rows[1:]:
        items.append({headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]})
    return items


def _parse_upload(filename: str, data: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return _parse_csv(data)
    if suffix == ".json":
        return _parse_json(data)
    if suffix in {".xlsx", ".xlsm"}:
        return _parse_xlsx(data)
    raise ValueError("Only CSV, JSON, XLSX and XLSM files are supported")


def build_metadata_template(format: str) -> bytes:
    normalized = format.lower().lstrip(".")
    if normalized not in TEMPLATE_FORMATS:
        raise ValueError("Only CSV, JSON, XLSX and XLSM templates are supported")
    if normalized == "csv":
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=TEMPLATE_HEADERS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(TEMPLATE_ROWS)
        return ("\ufeff" + output.getvalue()).encode("utf-8")
    if normalized == "json":
        return json.dumps({"items": TEMPLATE_ROWS}, ensure_ascii=False, indent=2).encode("utf-8")

    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ValueError("Excel template generation requires openpyxl to be installed") from exc
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "metadata"
    sheet.append(TEMPLATE_HEADERS)
    for row in TEMPLATE_ROWS:
        sheet.append([row.get(header, "") for header in TEMPLATE_HEADERS])
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 14), 36)
    output = BytesIO()
    workbook.save(output)
    data = output.getvalue()
    return _patch_xlsm_content_type(data) if normalized == "xlsm" else data


def _patch_xlsm_content_type(data: bytes) -> bytes:
    source = BytesIO(data)
    output = BytesIO()
    workbook_override = f"{{{CONTENT_TYPES_NS}}}Override"
    with zipfile.ZipFile(source, "r") as input_zip, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as output_zip:
        for item in input_zip.infolist():
            payload = input_zip.read(item.filename)
            if item.filename == CONTENT_TYPES_PATH:
                root = ElementTree.fromstring(payload)
                patched = False
                for override in root.findall(workbook_override):
                    if override.attrib.get("PartName") == "/xl/workbook.xml":
                        override.set("ContentType", XLSM_WORKBOOK_CONTENT_TYPE)
                        patched = True
                        break
                if not patched:
                    ElementTree.SubElement(
                        root,
                        workbook_override,
                        {"PartName": "/xl/workbook.xml", "ContentType": XLSM_WORKBOOK_CONTENT_TYPE},
                    )
                payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
            output_zip.writestr(item, payload)
    return output.getvalue()


def _find_work(db: Session, name: str) -> Work | None:
    return db.scalar(select(Work).where(Work.name == name))


def _find_character(db: Session, work_id: int, name: str) -> Character | None:
    return db.scalar(select(Character).where(Character.work_id == work_id, Character.name == name))


def _apply_fields(target, fields: dict[str, Any]) -> bool:
    changed = False
    for key, value in fields.items():
        if value is None:
            continue
        if getattr(target, key, None) != value:
            setattr(target, key, value)
            changed = True
    return changed


def _process_rows(db: Session, rows: list[dict[str, Any]], dry_run: bool) -> MetadataImportSummary:
    summary = MetadataImportSummary(dry_run=dry_run, total_rows=len(rows), valid_rows=0, error_rows=0)
    for index, raw in enumerate(rows, start=2):
        try:
            row = _normalize_row(raw, index)
            work_name = row["work_name"]
            character_name = row["character_name"]
            if not work_name:
                raise ValueError("work_name is required")

            work = _find_work(db, work_name)
            work_fields = {"name": work_name, **row["work"]}
            work_is_new = work is None
            row_work_created = 0
            row_work_updated = 0
            row_character_created = 0
            row_character_updated = 0
            if work:
                action = "update_work"
                if dry_run:
                    row_work_updated = 1
                elif _apply_fields(work, work_fields):
                    row_work_updated = 1
            else:
                action = "create_work"
                row_work_created = 1
                if not dry_run:
                    work = Work(**work_fields)
                    db.add(work)
                    db.flush()

            if character_name:
                character = None if work_is_new or not work else _find_character(db, work.id, character_name)
                character_fields = {**row["character"]}
                if work:
                    character_fields["work_id"] = work.id
                if character:
                    action = "update_character"
                    if dry_run:
                        row_character_updated = 1
                    elif _apply_fields(character, character_fields):
                        row_character_updated = 1
                else:
                    action = "create_character"
                    row_character_created = 1
                    if not dry_run:
                        if not work:
                            raise ValueError("work could not be created")
                        db.add(Character(**character_fields))

            summary.valid_rows += 1
            summary.works_created += row_work_created
            summary.works_updated += row_work_updated
            summary.characters_created += row_character_created
            summary.characters_updated += row_character_updated
            summary.rows.append(
                MetadataImportRow(
                    row_number=index,
                    action=action,
                    status="ok",
                    work_name=work_name,
                    character_name=character_name,
                    message="Ready" if dry_run else "Imported",
                )
            )
        except Exception as exc:
            summary.error_rows += 1
            summary.rows.append(
                MetadataImportRow(row_number=index, action="skip", status="error", message=str(exc))
            )
    if summary.error_rows and not dry_run:
        db.rollback()
    elif not dry_run:
        db.commit()
    return summary


@router.post("/metadata", response_model=MetadataImportSummary)
async def import_metadata(
    db: Annotated[Session, Depends(get_db)],
    admin: Annotated[dict, Depends(require_admin)],
    file: UploadFile = File(...),
    dry_run: bool = Query(True),
):
    try:
        data = await file.read()
        raw_rows = _parse_upload(file.filename or "", data)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return _process_rows(db, raw_rows, dry_run=dry_run)


@router.get("/metadata/template")
def download_metadata_template(
    admin: Annotated[dict, Depends(require_admin)],
    format: Literal["csv", "json", "xlsx", "xlsm"] = Query("xlsx"),
):
    try:
        data = build_metadata_template(format)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    filename = f"metadata-import-template.{format}"
    return Response(
        content=data,
        media_type=TEMPLATE_MEDIA_TYPES[format],
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"; filename*=UTF-8\'\'{quote(filename)}',
            "X-Template-Filename": filename,
        },
    )
